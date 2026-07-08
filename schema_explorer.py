#!/usr/bin/env python3
"""RCSB PDB Schema Explorer — Introspect the complete RCSB Data API schema.

Usage:
    python schema_explorer.py                    # Build everything
    python schema_explorer.py --search ligand    # Search for fields matching 'ligand'
    python schema_explorer.py --search binding   # Search for fields matching 'binding'
    python schema_explorer.py --statistics       # Print statistics only
    python schema_explorer.py --cross-reference  # Print cross-reference report only
    python schema_explorer.py --help             # Show help

Output files are written to the current directory.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from typing import Any

try:
    from rcsbapi.data import DataSchema
except ImportError:
    print("ERROR: rcsb-api package not found. Install with: pip install rcsb-api")
    sys.exit(1)


OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

CATEGORY_MAP = [
    (
        "Entry",
        lambda o, _: (
            o == "CoreEntry"
            or o == "Entry"
            or o == "CurrentEntry"
            or o == "StructKeywords"
            or o.startswith("RcsbEntry")
        ),
    ),
    (
        "Polymer Entity",
        lambda o, _: (
            (
                o == "CorePolymerEntity"
                or o.startswith("RcsbPolymerEntity")
                or o == "EntityPoly"
                or o == "GeneName"
                or o.startswith("RcsbEntitySource")
                or o.startswith("RcsbEntityHost")
                or o == "StructAsym"
            )
            and "Instance" not in o
        ),
    ),
    (
        "Polymer Entity Instance",
        lambda o, _: (
            "PolymerEntityInstance" in o
            or o == "CorePolymerEntityInstance"
            or o.startswith("RcsbPolymerInstance")
            or "PolymerInstanceFeature" in o
        ),
    ),
    ("Branched Entity", lambda o, _: o == "CoreBranchedEntity" or "BranchedEntity" in o or o == "PdbxEntityBranch"),
    (
        "Nonpolymer Entity",
        lambda o, _: (
            (o == "CoreNonpolymerEntity" or o.startswith("RcsbNonpolymerEntity") or o == "PdbxEntityNonpoly")
            and "Instance" not in o
            and "Annotation" not in o
        ),
    ),
    (
        "Nonpolymer Entity Instance",
        lambda o, _: "NonpolymerEntityInstance" in o or o.startswith("RcsbNonpolymerInstance"),
    ),
    (
        "Assembly",
        lambda o, _: (
            o == "CoreAssembly"
            or o.startswith("RcsbAssembly")
            or o.startswith("PdbxStructAssembly")
            or o == "PdbxStructOperList"
            or o == "PdbxStructSpecialSymmetry"
            or "StructSymmetry" in o
        ),
    ),
    ("Interface", lambda o, _: o == "CoreInterface" or o.startswith("RcsbInterface")),
    (
        "Chemical Component",
        lambda o, _: (
            o == "CoreChemComp"
            or o == "ChemComp"
            or o.startswith("RcsbChemComp")
            or o.startswith("PdbxChemComp")
            or o.startswith("PdbxReference")
        ),
    ),
    (
        "Experimental Data (X-ray)",
        lambda o, _: (
            o.startswith("Diffrn")
            or o in ("Cell", "Symmetry")
            or o == "Refine"
            or o.startswith("Refine")
            or o == "Reflns"
            or "Reflns" in o
            or o == "PdbxReflnsTwin"
        ),
    ),
    ("Experimental Data (EM)", lambda o, _: o.startswith("Em") or o.startswith("Em") or "Em" in o[:3]),
    ("Experimental Data (NMR)", lambda o, _: o.startswith("PdbxNmr")),
    ("Experimental Data (XFEL)", lambda o, _: o.startswith("PdbxSerialCrystallography") or o.startswith("PdbxSerial")),
    ("Experimental Data (SAXS)", lambda o, _: o.startswith("PdbxSolnScatter")),
    ("Validation Report", lambda o, _: o.startswith("PdbxVrpt")),
    ("Citation", lambda o, _: o == "Citation" or o == "RcsbPrimaryCitation" or o.startswith("RcsbBirdCitation")),
    ("PubMed Integration", lambda o, _: o == "CorePubmed" or o.startswith("RcsbPubmed") or "Pubmed" in o),
    (
        "UniProt Integration",
        lambda o, _: (
            o == "CoreUniprot" or o.startswith("RcsbUniprot") or "Uniprot" in o or o == "RcsbUniprotAlignments"
        ),
    ),
    ("DrugBank Integration", lambda o, _: o == "CoreDrugbank" or o.startswith("Drugbank") or "Drugbank" in o),
    ("Pfam Integration", lambda o, _: o == "CorePfam" or o.startswith("RcsbPfam") or "Pfam" in o),
    (
        "Target & Ligand Information",
        lambda o, _: (
            o.startswith("RcsbTarget")
            or o.startswith("RcsbLigand")
            or o.startswith("RcsbBinding")
            or "RcsbRelatedTarget" in o
        ),
    ),
    (
        "Computational Models",
        lambda o, _: (
            o.startswith("RcsbCompModel")
            or o.startswith("RcsbMaQaMetric")
            or o.startswith("RcsbIhm")
            or o.startswith("Ihm")
            or o == "MaData"
        ),
    ),
    (
        "Audit & Version History",
        lambda o, _: (
            o.startswith("PdbxAudit")
            or o.startswith("AuditAuthor")
            or o == "RcsbAccessionInfo"
            or o == "RcsbLatestRevision"
            or o == "CurrentEntry"
            or o.startswith("PdbxDatabase")
        ),
    ),
    ("Organism/Source", lambda o, _: o.startswith("EntitySrc") or o.startswith("PdbxEntitySrc")),
    (
        "Clusters & Groups",
        lambda o, _: (
            "ClustersMembers" in o
            or o.startswith("RcsbCluster")
            or o.startswith("RcsbGroup")
            or o.startswith("Group")
            or o.startswith("RcsbSchema")
        ),
    ),
    ("Software", lambda o, _: o == "Software" or "PdbxSoftware" in o),
    (
        "Database Cross-References",
        lambda o, _: o.startswith("RcsbExternalReferences") or o == "Database2" or o.startswith("PdbxDatabaseRelated"),
    ),
    ("Membrane & Genomic Context", lambda o, _: o.startswith("RcsbMembrane") or o.startswith("RcsbGenomic")),
    (
        "Sequence Features",
        lambda o, _: (
            "RcsbPolymerEntityAlign" in o
            or "RcsbPolymerEntityGroupSequenceAlignment" in o
            or "CoreEntityAlignments" in o
            or "AlignedRegion" in o
            or "Alignment" in o
            or "RcsbPolymerEntityContainerIdentifiersReferenceSequence" in o
        ),
    ),
]

DEFAULT_CATEGORY = "Other"


def resolve_type(type_dict: dict[str, Any]) -> tuple[str, str, bool, bool]:
    """Resolve GraphQL type information: (type_name, kind, nullable, is_list)."""
    kind = type_dict.get("kind", "")
    type_name = type_dict.get("name")
    nullable = kind != "NON_NULL"
    is_list = False
    oftype = type_dict.get("ofType")

    if kind in ("LIST", "NON_NULL") and oftype:
        sub_name, sub_kind, sub_nullable, sub_list = resolve_type(oftype)
        if not type_name or type_name == kind:
            type_name = sub_name
        if kind == "LIST" or sub_kind == "LIST":
            is_list = True
        return type_name, kind, nullable, is_list

    return type_name or "Unknown", kind, nullable, is_list


def get_enum_values(types_list: list[dict], type_name: str) -> list[str]:
    for t in types_list:
        if t.get("name") == type_name and t.get("kind") == "ENUM":
            return [v["name"] for v in t.get("enumValues") or []]
    return []


def find_category(object_name: str) -> str:
    for cat, matcher in CATEGORY_MAP:
        if matcher(object_name, ""):
            return cat
    return DEFAULT_CATEGORY


def is_builtin_type(type_name: str) -> bool:
    return type_name in ("String", "Int", "Float", "Boolean", "Date", "ObjectScalar")


def is_internal_type(type_name: str) -> bool:
    return type_name.startswith("__")


class SchemaExplorer:
    """Explore the RCSB Data API schema via GraphQL introspection."""

    def __init__(self) -> None:
        self.schema = DataSchema()
        self.raw = self.schema.schema["data"]["__schema"]
        self.types_list: list[dict] = self.raw["types"]

        self.type_map: dict[str, dict] = {}
        self.object_types: list[dict] = []
        self.enum_types: list[dict] = []
        self.scalar_types: set = set()

        for t in self.types_list:
            name = t.get("name", "")
            if is_internal_type(name):
                continue
            self.type_map[name] = t
            kind = t.get("kind")
            if kind == "OBJECT":
                self.object_types.append(t)
            elif kind == "ENUM":
                self.enum_types.append(t)
            elif kind == "SCALAR":
                self.scalar_types.add(name)

        for t in self.types_list:
            if t.get("name") == "__Type":
                # Some scalar types may not be directly listed
                pass

        self.query_type = self.type_map.get("Query")

    def get_root_fields(self) -> list[dict]:
        """Return the root query fields (entry, polymer_entities, etc.)."""
        if not self.query_type:
            return []
        root_fields = []
        for f in self.query_type.get("fields") or []:
            ret_type, kind, nullable, is_list = resolve_type(f["type"])
            root_fields.append(
                {
                    "name": f["name"],
                    "type": ret_type,
                    "kind": kind,
                    "nullable": nullable,
                    "is_list": is_list,
                    "description": f.get("description", "") or "",
                    "args": [{"name": a["name"]} for a in f.get("args") or []],
                }
            )
        return root_fields

    def get_object_fields(self, type_name: str) -> list[dict]:
        """Return all fields for a given object type."""
        t = self.type_map.get(type_name)
        if not t:
            return []
        fields = []
        for f in t.get("fields") or []:
            ret_type, kind, nullable, is_list = resolve_type(f["type"])
            enum_vals = get_enum_values(self.types_list, ret_type)
            fields.append(
                {
                    "name": f["name"],
                    "type": ret_type,
                    "kind": kind,
                    "nullable": nullable,
                    "is_list": is_list,
                    "description": f.get("description", "") or "",
                    "enum_values": enum_vals,
                }
            )
        return fields

    def list_all_fields_grouped_by_type(self) -> dict[str, list[dict]]:
        """Return a dict mapping each object type to its fields (direct fields only, no nested expansion)."""
        result: dict[str, list[dict]] = {}
        for t in self.object_types:
            name = t["name"]
            fields = self.get_object_fields(name)
            if fields:
                result[name] = fields
        return result

    def build_catalog(self) -> list[dict[str, Any]]:
        """Build a flat catalog: each field listed once per containing object type.

        This does NOT expand through shared sub-types to avoid combinatorial explosion.
        Instead, it lists each field once with its direct parent type.
        To find sub-fields, one can look up the referenced type in the catalog.
        """
        catalog = []
        root_fields = self.get_root_fields()

        type_fields_map = self.list_all_fields_grouped_by_type()

        for root in root_fields:
            root_name = root["name"]
            root_type = root["type"]

            root_obj_desc = self.type_map.get(root_type, {}).get("description", "") or ""

            fields = type_fields_map.get(root_type, [])
            for f in fields:
                catalog.append(
                    {
                        "object_name": root_type,
                        "object_description": root_obj_desc,
                        "parent_object": root_name,
                        "full_path": f"{root_name}.{f['name']}",
                        "field_name": f["name"],
                        "gql_type": f["type"],
                        "kind": f["kind"],
                        "nullable": f["nullable"],
                        "is_list": f["is_list"],
                        "description": f["description"],
                        "enum_values": f["enum_values"],
                        "category": find_category(root_type),
                        "sub_type": "object"
                        if f["type"] in type_fields_map
                        else ("enum" if f["type"] in [e["name"] for e in self.enum_types] else "scalar"),
                    }
                )

            if not is_builtin_type(root_type) and root_type in type_fields_map:
                sub_catalog = self._expand_type(root_type, root_name, type_fields_map, seen_types=set())
                catalog.extend(sub_catalog)

        return catalog

    def _expand_type(
        self,
        type_name: str,
        parent_path: str,
        type_fields_map: dict[str, list[dict]],
        seen_types: set,
    ) -> list[dict[str, Any]]:
        """Recursively expand fields of an object type once per parent path.

        Each type can appear multiple times in the catalog but only once per
        direct parent path. This avoids combinatorial explosion from shared
        sub-types while still showing all reachable fields.
        """
        if type_name in seen_types:
            return []
        seen_types.add(type_name)

        results = []
        fields = type_fields_map.get(type_name, [])
        obj_desc = self.type_map.get(type_name, {}).get("description", "") or ""

        for f in fields:
            full_path = f"{parent_path}.{f['name']}"
            results.append(
                {
                    "object_name": type_name,
                    "object_description": obj_desc,
                    "parent_object": parent_path.split(".")[-1],
                    "full_path": full_path,
                    "field_name": f["name"],
                    "gql_type": f["type"],
                    "kind": f["kind"],
                    "nullable": f["nullable"],
                    "is_list": f["is_list"],
                    "description": f["description"],
                    "enum_values": f["enum_values"],
                    "category": find_category(type_name),
                    "sub_type": "object"
                    if f["type"] in type_fields_map
                    else ("enum" if f["type"] in [e["name"] for e in self.enum_types] else "scalar"),
                }
            )

            if not is_builtin_type(f["type"]) and f["type"] in type_fields_map:
                sub = self._expand_type(f["type"], full_path, type_fields_map, seen_types.copy())
                results.extend(sub)

        return results

    def search(self, query: str) -> list[dict[str, Any]]:
        """Search the catalog for matching field names, object names, descriptions, or paths."""
        types_map = self.list_all_fields_grouped_by_type()
        flattened = []
        for type_name, fields in types_map.items():
            obj_desc = self.type_map.get(type_name, {}).get("description", "") or ""
            for f in fields:
                flattened.append(
                    {
                        "object_name": type_name,
                        "object_description": obj_desc,
                        "field_name": f["name"],
                        "full_path": f"{type_name}.{f['name']}",
                        "gql_type": f["type"],
                        "nullable": f["nullable"],
                        "is_list": f["is_list"],
                        "description": f["description"],
                        "enum_values": f["enum_values"],
                        "category": find_category(type_name),
                    }
                )

        q = query.lower()
        results = []
        for rec in flattened:
            if (
                q in rec["object_name"].lower()
                or q in rec["field_name"].lower()
                or q in (rec["description"] or "").lower()
                or q in rec["full_path"].lower()
            ):
                results.append(rec)
        return results

    def get_statistics(self) -> dict[str, Any]:
        """Compute summary statistics about the schema."""
        type_fields_map = self.list_all_fields_grouped_by_type()

        total_fields = sum(len(fields) for fields in type_fields_map.values())

        object_field_counts: dict[str, int] = {name: len(fields) for name, fields in type_fields_map.items()}

        total_scalar = 0
        total_enum_fields = 0
        total_object_fields = 0
        for _name, fields in type_fields_map.items():
            for f in fields:
                if is_builtin_type(f["type"]):
                    total_scalar += 1
                elif f["type"] in [e["name"] for e in self.enum_types]:
                    total_enum_fields += 1
                elif f["type"] in type_fields_map:
                    total_object_fields += 1

        largest_obj = max(object_field_counts, key=lambda k: object_field_counts[k])
        largest_obj_count = object_field_counts[largest_obj]

        categories: dict[str, int] = defaultdict(int)
        for name in type_fields_map:
            cat = find_category(name)
            categories[cat] += len(type_fields_map[name])

        return {
            "total_object_types": len(self.object_types),
            "total_enum_types": len(self.enum_types),
            "total_fields": total_fields,
            "scalar_fields": total_scalar,
            "enum_fields": total_enum_fields,
            "object_reference_fields": total_object_fields,
            "deepest_nesting": self._compute_max_nesting(type_fields_map),
            "largest_object": largest_obj,
            "largest_object_field_count": largest_obj_count,
            "object_names": sorted(type_fields_map.keys()),
            "categories": dict(categories),
            "object_field_counts": {k: v for k, v in sorted(object_field_counts.items(), key=lambda x: -x[1])},
        }

    def _compute_max_nesting(self, type_fields_map: dict[str, list[dict]]) -> int:
        """Determine deepest nesting level by tracing object references."""

        def depth(type_name: str, visited: set) -> int:
            if type_name in visited or type_name not in type_fields_map:
                return 0
            visited.add(type_name)
            max_d = 0
            for f in type_fields_map.get(type_name, []):
                if not is_builtin_type(f["type"]) and f["type"] in type_fields_map:
                    d = 1 + depth(f["type"], visited.copy())
                    max_d = max(max_d, d)
            return max_d

        max_nest = 0
        for type_name in type_fields_map:
            d = depth(type_name, set())
            max_nest = max(max_nest, d)
        return max_nest

    def get_cross_reference_data(self) -> dict[str, list[dict]]:
        """Find fields related to cross-reference databases."""
        topics: dict[str, list[str]] = {
            "UniProt": ["uniprot", "uni_prot", "uniprotkb", "uniprot_primary", "uniprot_id"],
            "DrugBank": ["drugbank", "drug_bank", "drugbank_id"],
            "PubMed": ["pubmed", "pubmed_id", "pmid", "pmc"],
            "GO (Gene Ontology)": ["go_", "gene_ontology", "go_id", "go_term", "go_aspect", "go_evidence"],
            "EC (Enzyme Classification)": [
                "ec_",
                "enzyme_class",
                "enzyme_commission",
                "ec_number",
                "rcsb_enzyme_class",
            ],
            "Pfam": ["pfam", "pfam_id", "pfam_acc", "pfam_classification"],
            "SCOP": ["scop", "scop_id", "scop_classification", "scopedb"],
            "CATH": ["cath", "cath_id", "cath_classification", "cathdb"],
            "Ligands": ["ligand", "ligand_id", "chem_comp", "nonpolymer", "non_polymer", "chemical_component"],
            "Binding Sites": ["binding", "bind", "binding_site", "binding_affinity", "target_neighbor"],
            "Active Sites": ["active_site", "active", "catalytic"],
            "Residues": ["residue", "resid", "res_name", "comp_id", "monomer", "auth_comp_id", "mon_id"],
            "Chains": ["chain", "asym_id", "auth_asym_id", "polymer_entity_instance", "struct_asym", "pdbx_strand_id"],
            "Assemblies": ["assembly", "assemblies", "oligomeric", "oligomeric_state", "pdbx_struct_assembly"],
            "Organisms": [
                "organism",
                "organism_name",
                "organism_scientific",
                "source_organism",
                "host_organism",
                "taxonomy",
                "tax_id",
                "ncbi",
            ],
            "Sequence Features": [
                "sequence",
                "seq_one_letter",
                "seq_",
                "entity_poly",
                "poly_seq",
                "alignment",
                "aligned_region",
                "sequence_identity",
                "seq_length",
            ],
            "Membrane Proteins": ["membrane", "transmembrane", "membrane_lineage"],
            "Genomic": ["genomic", "genome", "gene_name", "genomic_lineage", "gene"],
            "Keywords": ["keyword", "keywords"],
            "DOI / Citation": ["doi", "citation", "journal", "journal_abbrev"],
            "Validation Metrics": [
                "validation",
                "clashscore",
                "clash_score",
                "rama",
                "ramachandran",
                "rmsd",
                "rms",
                "molprobity",
            ],
            "Symmetry": ["symmetry", "space_group", "crystal"],
        }

        types_map = self.list_all_fields_grouped_by_type()
        all_fields = []
        for type_name, fields in types_map.items():
            for f in fields:
                all_fields.append(
                    {
                        "object_name": type_name,
                        "field_name": f["name"],
                        "full_path": f"{type_name}.{f['name']}",
                        "gql_type": f["type"],
                        "nullable": f["nullable"],
                        "is_list": f["is_list"],
                        "description": f["description"],
                        "category": find_category(type_name),
                    }
                )

        results: dict[str, list[dict]] = {}
        for topic, patterns in topics.items():
            matches = []
            for rec in all_fields:
                combined = f"{rec['object_name']} {rec['field_name']} {rec['description']} {rec['full_path']}".lower()
                if any(p.lower() in combined for p in patterns):
                    matches.append(rec)
            if matches:
                results[topic] = matches

        return results


# ─── Export Functions ─────────────────────────────────────────────────


def export_json(catalog: list[dict[str, Any]], filepath: str) -> None:
    with open(filepath, "w") as f:
        json.dump(catalog, f, indent=2, default=str)
    print(f"  Wrote {filepath} ({len(catalog)} records)")


def export_csv(catalog: list[dict[str, Any]], filepath: str) -> None:
    if not catalog:
        return
    fieldnames = [
        "object_name",
        "object_description",
        "parent_object",
        "full_path",
        "field_name",
        "gql_type",
        "kind",
        "nullable",
        "is_list",
        "description",
        "enum_values",
        "category",
        "sub_type",
    ]
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in catalog:
            r = {k: row.get(k, "") for k in fieldnames}
            if isinstance(r["enum_values"], list):
                r["enum_values"] = "; ".join(r["enum_values"])
            writer.writerow(r)
    print(f"  Wrote {filepath} ({len(catalog)} records)")


def export_markdown(catalog: list[dict[str, Any]], filepath: str, stats: dict[str, Any]) -> None:
    type_fields: dict[str, list[dict]] = defaultdict(list)
    for r in catalog:
        type_fields[r["object_name"]].append(r)

    lines = [
        "# RCSB PDB Schema Catalogue\n",
        f"_Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} from live RCSB Data API_\n",
        f"**Total Object Types:** {stats['total_object_types']}  \n",
        f"**Total Fields:** {stats['total_fields']}  \n",
        f"**Deepest Nesting:** {stats['deepest_nesting']} levels  \n",
        f"**Largest Object:** `{stats['largest_object']}` ({stats['largest_object_field_count']} fields)\n",
        "---\n",
    ]

    for obj_name in sorted(type_fields.keys()):
        fields = sorted(type_fields[obj_name], key=lambda x: x["field_name"])
        obj_desc = fields[0]["object_description"]
        cat = fields[0].get("category", DEFAULT_CATEGORY)
        lines.append(f"\n## {obj_name}\n")
        if obj_desc:
            lines.append(f"_{obj_desc}_\n")
        lines.append(f"**Category:** {cat}  \n" if cat != DEFAULT_CATEGORY else "")
        lines.append("| Field | Type | Nullable | List | Description |\n")
        lines.append("|-------|------|----------|------|-------------|\n")
        for fd in fields:
            null_s = "Yes" if fd["nullable"] else "No"
            list_s = "Yes" if fd["is_list"] else "No"
            desc = (fd["description"] or "").replace("\n", " ").replace("|", "\\|")[:120]
            enum_s = ""
            if fd["enum_values"]:
                enum_s = "(" + ", ".join(fd["enum_values"][:5]) + ")"
                if len(fd["enum_values"]) > 5:
                    enum_s += "..."
            gtype = fd["gql_type"]
            if enum_s:
                gtype = f"ENUM {enum_s}"
            lines.append(f"| `{fd['field_name']}` | {gtype} | {null_s} | {list_s} | {desc} |\n")

    with open(filepath, "w") as fh:
        fh.writelines(lines)
    print(f"  Wrote {filepath}")


def export_xlsx(catalog: list[dict[str, Any]], filepath: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("  Skipping .xlsx: openpyxl not installed (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RCSB Schema Fields"

    fieldnames = [
        ("object_name", "Object Name"),
        ("object_description", "Object Description"),
        ("parent_object", "Parent Object"),
        ("full_path", "Full Path"),
        ("field_name", "Field Name"),
        ("gql_type", "GraphQL Type"),
        ("kind", "Kind"),
        ("nullable", "Nullable"),
        ("is_list", "Is List"),
        ("description", "Description"),
        ("enum_values", "Enum Values"),
        ("category", "Category"),
    ]

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col, (_, label) in enumerate(fieldnames, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hf
        c.fill = hfill

    for ri, rec in enumerate(catalog, 2):
        for ci, (key, _) in enumerate(fieldnames, 1):
            val = rec.get(key, "")
            if key == "enum_values" and isinstance(val, list):
                val = "; ".join(val)
            ws.cell(row=ri, column=ci, value=val)

    ws.auto_filter.ref = f"A1:{chr(64 + len(fieldnames))}{len(catalog) + 1}"
    wb.save(filepath)
    print(f"  Wrote {filepath} ({len(catalog)} records)")


def export_statistics(stats: dict[str, Any], filepath: str) -> None:
    lines = [
        "# RCSB PDB Schema Statistics\n",
        f"_Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_\n",
        "## Summary\n",
        f"- **Total Object Types (GraphQL types):** {stats['total_object_types']}\n",
        f"- **Total Enum Types:** {stats['total_enum_types']}\n",
        f"- **Total Fields:** {stats['total_fields']}\n",
        f"  - Scalar fields: {stats['scalar_fields']}\n",
        f"  - Enum field references: {stats['enum_fields']}\n",
        f"  - Object reference fields: {stats['object_reference_fields']}\n",
        f"- **Deepest Nesting Level:** {stats['deepest_nesting']}\n",
        f"- **Largest Object:** `{stats['largest_object']}` ({stats['largest_object_field_count']} fields)\n",
        "\n## Objects by Field Count\n\n",
        "| Object | Fields |\n",
        "|--------|-------:|\n",
    ]
    for obj_name, count in stats["object_field_counts"].items():
        lines.append(f"| `{obj_name}` | {count} |\n")

    lines.extend(
        [
            "\n## Fields by Category\n\n",
            "| Category | Fields |\n",
            "|----------|-------:|\n",
        ]
    )
    for cat, count in sorted(stats["categories"].items(), key=lambda x: -x[1]):
        lines.append(f"| {cat} | {count} |\n")

    lines.extend(
        [
            "\n## All Object Names\n",
            f"*{len(stats['object_names'])} object types*\n",
        ]
    )
    for name in stats["object_names"]:
        lines.append(f"- `{name}`\n")

    with open(filepath, "w") as f:
        f.writelines(lines)
    print(f"  Wrote {filepath}")


def export_cross_reference_md(results: dict[str, list[dict]], filepath: str) -> None:
    all_count = sum(len(v) for v in results.values())
    lines = [
        "# Cross-Reference Report\n",
        f"_Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_\n",
        "Fields related to cross-reference databases and biological concepts.\n",
        f"**Total related field references:** {all_count}\n",
        f"**Topics covered:** {len(results)}\n",
    ]

    for topic in sorted(results.keys()):
        matches = sorted(results[topic], key=lambda x: x["full_path"])
        lines.extend(
            [
                f"\n## {topic} ({len(matches)} fields)\n\n",
                "| Full Path | Field | Object | Type | Description |\n",
                "|-----------|-------|--------|------|-------------|\n",
            ]
        )
        for rec in matches:
            desc = (rec["description"] or "").replace("\n", " ")[:100]
            lines.append(
                f"| `{rec['full_path']}` | `{rec['field_name']}` | {rec['object_name']} | {rec['gql_type']} | {desc} |\n"
            )

    with open(filepath, "w") as f:
        f.writelines(lines)
    print(f"  Wrote {filepath}")


def export_hierarchy_md(explorer: SchemaExplorer, filepath: str) -> None:
    """Export a compact type hierarchy showing each object type and its direct fields."""
    type_fields_map = explorer.list_all_fields_grouped_by_type()
    root_fields = explorer.get_root_fields()

    lines = [
        "# RCSB Field Catalogue (Type Hierarchy)\n",
        f"_Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_\n",
        "This shows each GraphQL object type and its direct fields.\n",
        "Object reference fields show the nested type name in brackets.\n",
    ]

    root_to_type: dict[str, list[str]] = {}
    for rf in root_fields:
        root_to_type[rf["type"]] = root_to_type.get(rf["type"], []) + [rf["name"]]

    for type_name in sorted(type_fields_map.keys()):
        fields = sorted(type_fields_map[type_name], key=lambda x: x["name"])
        roots = root_to_type.get(type_name, [])
        if roots:
            rstr = ", ".join(roots)
            lines.append(f"\n## {type_name}  (from root query: {rstr})\n")
        else:
            lines.append(f"\n## {type_name}\n")

        lines.append("```text\n")
        for fd in fields:
            gtype = fd["type"]
            if fd["enum_values"]:
                gtype = "ENUM"
            null_flag = "?" if fd["nullable"] else ""
            list_flag = "[]" if fd["is_list"] else ""
            dtype = f"{gtype}{list_flag}{null_flag}"
            desc = (fd["description"] or "").replace("\n", " ").strip()[:120]
            if desc:
                lines.append(f"  ├── {fd['name']} : {dtype}  # {desc}\n")
            else:
                lines.append(f"  ├── {fd['name']} : {dtype}\n")
        lines.append("```\n")

    with open(filepath, "w") as fh:
        fh.writelines(lines)
    print(f"  Wrote {filepath}")


def print_search_results(results: list[dict], query: str) -> None:
    print(f"\n=== Search results for '{query}': {len(results)} fields ===\n")
    if not results:
        print("  No results found.\n")
        return
    print(f"{'Object':<30} {'Field':<30} {'Type':<20} {'Description'}")
    print(f"{'-' * 30} {'-' * 30} {'-' * 20} {'-' * 60}")
    for r in results[:60]:
        desc = (r["description"] or "")[:60]
        print(f"{r['object_name']:<30} {r['field_name']:<30} {r['gql_type']:<20} {desc}")
    if len(results) > 60:
        print(f"  ... and {len(results) - 60} more results")


# ─── CLI ──────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="RCSB PDB Schema Explorer — Introspect the complete RCSB Data API schema"
    )
    parser.add_argument("--search", type=str, help="Search for fields matching a query")
    parser.add_argument("--statistics", action="store_true", help="Print statistics only (to stdout)")
    parser.add_argument("--cross-reference", action="store_true", help="Print cross-reference report only (to stdout)")
    parser.add_argument("--no-export", action="store_true", help="Skip writing output files")
    parser.add_argument("--json", default="", help="Custom JSON output path")
    parser.add_argument("--csv", default="", help="Custom CSV output path")
    parser.add_argument("--xlsx", default="", help="Custom XLSX output path")
    parser.add_argument("--md", default="", help="Custom Markdown output path")
    args = parser.parse_args()

    if args.search:
        explorer = SchemaExplorer()
        results = explorer.search(args.search)
        print_search_results(results, args.search)
        return

    print("Exploring RCSB Data API schema...")
    print("(This queries the live RCSB GraphQL introspection endpoint)\n")

    explorer = SchemaExplorer()
    print(f"Found {len(explorer.object_types)} object types, {len(explorer.enum_types)} enum types\n")

    stats = explorer.get_statistics()
    print("=== Schema Statistics ===")
    print(f"  Object Types:         {stats['total_object_types']}")
    print(f"  Enum Types:           {stats['total_enum_types']}")
    print(f"  Total Fields:         {stats['total_fields']}")
    print(f"  Scalar Fields:        {stats['scalar_fields']}")
    print(f"  Enum Field Refs:      {stats['enum_fields']}")
    print(f"  Object Refs:          {stats['object_reference_fields']}")
    print(f"  Deepest Nesting:      {stats['deepest_nesting']} levels")
    print(f"  Largest Object:       {stats['largest_object']} ({stats['largest_object_field_count']} fields)\n")

    if args.statistics:
        print(json.dumps(stats, indent=2, default=str))
        return

    catalog = explorer.build_catalog()
    print(f"Catalog records (unique type.field entries): {len(catalog)}\n")

    if args.cross_reference:
        xref = explorer.get_cross_reference_data()
        print(json.dumps({k: len(v) for k, v in sorted(xref.items())}, indent=2))
        return

    if args.no_export:
        print("Skipping file exports (--no-export)")
        return

    print("Exporting...")
    export_json(catalog, args.json or os.path.join(OUTPUT_DIR, "rcsb_schema.json"))
    export_csv(catalog, args.csv or os.path.join(OUTPUT_DIR, "rcsb_schema.csv"))
    export_markdown(catalog, args.md or os.path.join(OUTPUT_DIR, "rcsb_schema.md"), stats)
    export_xlsx(catalog, args.xlsx or os.path.join(OUTPUT_DIR, "rcsb_schema.xlsx"))
    export_hierarchy_md(explorer, os.path.join(OUTPUT_DIR, "rcsb_schema_hierarchy.md"))
    export_statistics(stats, os.path.join(OUTPUT_DIR, "field_statistics.md"))

    xref = explorer.get_cross_reference_data()
    export_cross_reference_md(xref, os.path.join(OUTPUT_DIR, "cross_reference_report.md"))

    print("\nDone! All files written.")


if __name__ == "__main__":
    main()
